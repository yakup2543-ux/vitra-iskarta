import streamlit as st
import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Iskarta Sistemi", layout="wide")

st.title("🏭 VitrA Iskarta Kayıt Mobil Uygulaması")

# Ekranı iki sekmeye ayırıyoruz
sekme1, sekme2 = st.tabs(["📝 Veri Girişi", "📊 Canlı Üretim Analizi"])

# ==========================================
# 1. SEKME: SENİN ORİJİNAL VERİ GİRİŞ MOTORUN
# ==========================================
with sekme1:
    tarih = st.text_input("Tarih", datetime.now().strftime("%d.%m.%Y %H:%M"))
    hat = st.text_input("Hat")
    urun = st.text_input("Ürün İsmi")
    hata = st.text_input("Hata Adı")
    neden = st.text_input("Muhtemel Neden")
    sonuc = st.text_input("Sonuç")
    notlar = st.text_area("Notlar")

    uploaded_files = st.file_uploader("Hata Fotoğrafları (Maks 15)", accept_multiple_files=True, type=['jpg', 'jpeg', 'png'])

    if st.button("Veriyi Kaydet (Excel'e Aktar)"):
        if not hat or not hata:
            st.error("Hat ve Hata Adı boş bırakılamaz!")
        else:
            DOSYA_ADI = "Iskarta_Bilgi_Sistemi.xlsx"
            
            if not os.path.exists("Temp_Fotolar"):
                os.makedirs("Temp_Fotolar")

            try:
                if not os.path.exists(DOSYA_ADI):
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Iskarta Kayıtları"
                    
                    basliklar = ["Tarih", "Hat", "Ürün", "Hata Adı", "Muhtemel Neden", "Sonuç", "Not"]
                    for i in range(1, 16):
                        basliklar.append(f"Fotoğraf {i}")
                        
                    for col_idx, baslik in enumerate(basliklar, 1):
                        hucre = ws.cell(row=1, column=col_idx, value=baslik)
                        hucre.font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
                        hucre.fill = PatternFill(start_color="0052cc", end_color="0052cc", fill_type="solid")
                        hucre.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[1].height = 25
                else:
                    wb = openpyxl.load_workbook(DOSYA_ADI)
                    ws = wb.active

                satir_no = ws.max_row + 1
                yeni_veri = [tarih, hat, urun, hata, neden, sonuc, notlar]
                for col_idx, deger in enumerate(yeni_veri, 1):
                    ws.cell(row=satir_no, column=col_idx, value=deger)
                
                if uploaded_files:
                    fotolar = uploaded_files[:15]
                    ws.row_dimensions[satir_no].height = 95
                    
                    for indeks, foto in enumerate(fotolar):
                        sutun_no = 8 + indeks
                        sutun_harfi = get_column_letter(sutun_no)
                        
                        foto_yolu = os.path.join("Temp_Fotolar", foto.name)
                        with open(foto_yolu, "wb") as f:
                            f.write(foto.getbuffer())
                            
                        img = ExcelImage(foto_yolu)
                        img.width = 120
                        img.height = 120
                        ws.column_dimensions[sutun_harfi].width = 18
                        ws.add_image(img, f"{sutun_harfi}{satir_no}")
                
                wb.save(DOSYA_ADI)
                st.success(f"Başarılı! Veriler masaüstündeki '{DOSYA_ADI}' dosyasına eklendi.")
                
            except Exception as e:
                st.error(f"Excel'e yazarken bir hata oluştu: {e}")

# ==========================================
# 2. SEKME: YENİ CANLI ANALİZ VE GRAFİKLER
# ==========================================
with sekme2:
    st.header("📈 Üretim Hattı Canlı Analiz Paneli")
    DOSYA_ADI = "Iskarta_Bilgi_Sistemi.xlsx"
    
    if os.path.exists(DOSYA_ADI):
        try:
            # Sadece veri sütunlarını okuyoruz (fotoğraf sütunlarını analize dahil etmiyoruz)
            df = pd.read_excel(DOSYA_ADI, usecols=["Tarih", "Hat", "Ürün", "Hata Adı", "Muhtemel Neden", "Sonuç", "Not"])
            
            if df.empty:
                st.info("Excel dosyası oluşturulmuş ama henüz veri girilmemiş.")
            else:
                st.metric(label="Bugüne Kadarki Toplam Iskarta Kayıt Sayısı", value=len(df))
                st.markdown("---")
                
                # Grafikleri yan yana göstermek için ekranı iki kolona ayırıyoruz
                kolon1, kolon2 = st.columns(2)
                
                with kolon1:
                    st.subheader("Hangi Hatta Ne Kadar Hata Çıktı?")
                    hat_sayilari = df["Hat"].value_counts()
                    st.bar_chart(hat_sayilari)
                    
                with kolon2:
                    st.subheader("En Sık Görülen Hata Türleri")
                    hata_sayilari = df["Hata Adı"].value_counts()
                    st.bar_chart(hata_sayilari)
                    
                st.markdown("---")
                st.subheader("Son Girilen 10 Iskarta Kaydı")
                st.dataframe(df.tail(10).iloc[::-1], use_container_width=True)
                
        except Exception as e:
            st.error(f"Grafikler oluşturulurken bir hata oluştu: {e}")
            st.info("Not: Pandas kütüphanesi eksik olabilir. Sistemi başlatmadan önce CMD'ye 'pip install pandas' yazmanız gerekebilir.")
    else:
        st.warning("Henüz Excel dosyası oluşturulmamış. Lütfen 'Veri Girişi' sekmesinden ilk kaydınızı yapın.")